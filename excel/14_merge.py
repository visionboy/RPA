from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells("B2:D2") # B2 부터 D2 까지 합치겠음
ws["B2"].value = "Merged Cell"
cel = ws["B2"]
cel.alignment = Alignment(horizontal="center", vertical="center")
wb.save("sample_merge.xlsx")